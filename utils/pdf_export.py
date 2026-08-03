import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from database import db_session
from models import Customer, Technician, JobCard, JobCardTask, PartUsed, TimeEntry, Setting


def _get_company():
    def get(key, default=""):
        s = db_session.query(Setting).filter(Setting.key == key).first()
        return s.value if s and s.value else default
    return {
        "name": get("company_name", "JobCard System"),
        "address": get("company_address", "123 Business Park"),
        "phone": get("company_phone", "+27 11 000 0000"),
        "email": get("company_email", "info@jobcardsystem.co.za"),
    }


def generate_jobcard_print_html(jobcard):
    co = _get_company()
    now = datetime.now().strftime("%Y-%m-%d %H:%M")
    tasks = db_session.query(JobCardTask).filter(JobCardTask.jobcard_id == jobcard.id).order_by(JobCardTask.sort_order).all()
    parts = db_session.query(PartUsed).filter(PartUsed.jobcard_id == jobcard.id).all()
    times = db_session.query(TimeEntry).filter(TimeEntry.jobcard_id == jobcard.id).all()
    customer = db_session.query(Customer).get(jobcard.customer_id)
    tech = db_session.query(Technician).get(jobcard.technician_id) if jobcard.technician_id else None

    html = f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
body {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 12px; color: #222; margin: 20px; background: #fff; }}
.header {{ display: flex; justify-content: space-between; border-bottom: 2px solid #0d6efd; padding-bottom: 10px; margin-bottom: 15px; }}
.header h1 {{ margin: 0; font-size: 20px; color: #0d6efd; }}
.header .doc-info {{ text-align: right; }}
.header .doc-info h2 {{ margin: 0; font-size: 16px; color: #333; }}
.meta {{ margin-bottom: 15px; display: flex; gap: 30px; flex-wrap: wrap; }}
.meta .label {{ font-size: 10px; color: #666; text-transform: uppercase; }}
.meta .value {{ font-size: 13px; font-weight: 600; }}
.section {{ margin-bottom: 15px; }}
.section h3 {{ font-size: 14px; color: #0d6efd; border-bottom: 1px solid #ddd; padding-bottom: 5px; }}
table {{ width: 100%; border-collapse: collapse; margin-bottom: 10px; }}
th {{ background: #f0f0f0; padding: 6px; text-align: left; font-size: 10px; text-transform: uppercase; border-bottom: 2px solid #ccc; }}
td {{ padding: 5px; border-bottom: 1px solid #eee; font-size: 12px; }}
.status-badge {{ display: inline-block; padding: 2px 8px; border-radius: 10px; font-size: 11px; font-weight: 600; }}
.status-open {{ background: #e3f2fd; color: #1565c0; }}
.status-assigned {{ background: #fff3e0; color: #e65100; }}
.status-in_progress {{ background: #e8f5e9; color: #2e7d32; }}
.status-completed {{ background: #e8f5e9; color: #1b5e20; }}
.status-closed {{ background: #f3e5f5; color: #6a1b9a; }}
.status-cancelled {{ background: #fbe9e7; color: #bf360c; }}
.priority-critical {{ background: #ffebee; color: #c62828; }}
.priority-high {{ background: #fff3e0; color: #e65100; }}
.priority-medium {{ background: #fff8e1; color: #f9a825; }}
.priority-low {{ background: #e8f5e9; color: #2e7d32; }}
.footer {{ margin-top: 30px; border-top: 1px solid #ccc; padding-top: 10px; font-size: 10px; color: #666; text-align: center; }}
.notes-box {{ margin: 10px 0; padding: 10px; background: #f9f9f9; border-left: 3px solid #0d6efd; }}
.signature-box {{ margin-top: 30px; display: flex; justify-content: space-between; }}
.signature-line {{ border-top: 1px solid #333; width: 200px; padding-top: 5px; font-size: 11px; text-align: center; }}
</style></head><body>
<div class="header">
  <div><h1>{co['name']}</h1><div style="font-size:11px;color:#666;">{co['address']}<br>{co['phone']} | {co['email']}</div></div>
  <div class="doc-info"><h2>Job Card</h2><div style="font-size:12px;color:#666;">{jobcard.job_number}<br>{now}</div></div>
</div>

<div class="section">
  <h3>Job Information</h3>
  <div class="meta">
    <div><div class="label">Status</div><div class="value"><span class="status-badge status-{jobcard.status.value}">{jobcard.status.value.replace('_',' ').title()}</span></div></div>
    <div><div class="label">Priority</div><div class="value"><span class="status-badge priority-{jobcard.priority.value}">{jobcard.priority.value.upper()}</span></div></div>
    <div><div class="label">Title</div><div class="value">{jobcard.title}</div></div>
    <div><div class="label">Due Date</div><div class="value">{jobcard.due_date or 'N/A'}</div></div>
    <div><div class="label">Scheduled</div><div class="value">{jobcard.scheduled_date or 'N/A'}</div></div>
    <div><div class="label">Completed</div><div class="value">{jobcard.completed_date or 'N/A'}</div></div>
  </div>
</div>

<div class="section">
  <h3>Customer</h3>
  <div class="meta">
    <div><div class="label">Name</div><div class="value">{customer.name if customer else 'N/A'}</div></div>
    <div><div class="label">Company</div><div class="value">{customer.company if customer else ''}</div></div>
    <div><div class="label">Contact</div><div class="value">{customer.contact_person if customer else ''}</div></div>
    <div><div class="label">Phone</div><div class="value">{customer.phone if customer else ''}</div></div>
    <div><div class="label">Email</div><div class="value">{customer.email if customer else ''}</div></div>
    <div><div class="label">Address</div><div class="value">{customer.address if customer else ''}</div></div>
  </div>
</div>

<div class="section">
  <h3>Site & Assignment</h3>
  <div class="meta">
    <div><div class="label">Site Address</div><div class="value">{jobcard.site_address or 'N/A'}</div></div>
    <div><div class="label">Site Contact</div><div class="value">{jobcard.site_contact or 'N/A'}</div></div>
    <div><div class="label">Technician</div><div class="value">{tech.name if tech else 'Unassigned'}</div></div>
    <div><div class="label">Customer PO</div><div class="value">{jobcard.customer_po or 'N/A'}</div></div>
  </div>
</div>

<div class="section">
  <h3>Description</h3>
  <p>{jobcard.description or 'No description provided'}</p>
</div>"""

    if tasks:
        html += """<div class="section"><h3>Task List</h3><table>
<thead><tr><th>#</th><th>Description</th><th>Status</th><th>Est. Minutes</th><th>Actual Minutes</th></tr></thead><tbody>"""
        for i, t in enumerate(tasks, 1):
            html += f"<tr><td>{i}</td><td>{t.description}</td><td>{t.status.value.title()}</td><td>{t.estimated_minutes or 0}</td><td>{t.actual_minutes or 0}</td></tr>"
        html += "</tbody></table></div>"

    if parts:
        html += """<div class="section"><h3>Parts Used</h3><table>
<thead><tr><th>Part</th><th>SKU</th><th>Qty</th><th>Unit Cost</th><th>Total</th></tr></thead><tbody>"""
        for p in parts:
            html += f"<tr><td>{p.part_name}</td><td>{p.part_sku or '-'}</td><td>{p.quantity}</td><td>R{p.unit_cost:,.2f}</td><td>R{p.total_cost:,.2f}</td></tr>"
        html += "</tbody></table></div>"

    if times:
        html += """<div class="section"><h3>Time Entries</h3><table>
<thead><tr><th>Technician</th><th>Start</th><th>End</th><th>Hours</th><th>Rate</th><th>Cost</th></tr></thead><tbody>"""
        for t in times:
            html += f"<tr><td>{t.technician_name or ''}</td><td>{t.start_time.strftime('%Y-%m-%d %H:%M') if t.start_time else ''}</td><td>{t.end_time.strftime('%Y-%m-%d %H:%M') if t.end_time else ''}</td><td>{t.hours}</td><td>R{t.hourly_rate:,.2f}</td><td>R{t.total_cost:,.2f}</td></tr>"
        html += "</tbody></table></div>"

    html += f"""<div class="section"><h3>Cost Summary</h3>
<div class="meta">
  <div><div class="label">Parts Cost</div><div class="value">R{jobcard.total_parts_cost:,.2f}</div></div>
  <div><div class="label">Labour Cost</div><div class="value">R{jobcard.total_labour_cost:,.2f}</div></div>
  <div><div class="label">Total Cost</div><div class="value">R{jobcard.total_cost:,.2f}</div></div>
  <div><div class="label">Amount Charged</div><div class="value">R{jobcard.amount_charged:,.2f}</div></div>
  <div><div class="label">Est. Hours</div><div class="value">{jobcard.estimated_hours}</div></div>
  <div><div class="label">Actual Hours</div><div class="value">{jobcard.actual_hours}</div></div>
</div></div>"""

    if jobcard.resolution_notes:
        html += f'<div class="section"><h3>Resolution Notes</h3><div class="notes-box">{jobcard.resolution_notes}</div></div>'

    if jobcard.customer_notes:
        html += f'<div class="section"><h3>Customer Notes</h3><div class="notes-box">{jobcard.customer_notes}</div></div>'

    customer_sig = '<div class="signature-line">Customer Signature & Date</div>'
    if jobcard.signed_off and jobcard.signature_data:
        sig_when = jobcard.signed_off_at.strftime("%Y-%m-%d %H:%M") if jobcard.signed_off_at else ""
        customer_sig = (
            '<div style="text-align:center;">'
            f'<img src="{jobcard.signature_data}" '
            'style="max-height:70px;max-width:220px;background:#fff;border:1px solid #ddd;padding:4px;border-radius:4px;">'
            f'<div style="font-size:11px;">{jobcard.signed_off_by or "Customer"}</div>'
            f'<div style="font-size:10px;color:#666;">{sig_when}</div>'
            '</div>'
        )

    html += f"""<div class="signature-box">
  <div class="signature-line">Technician Signature & Date</div>
  {customer_sig}
</div>

<div class="footer">
  {co['name']} | {co['phone']} | {co['email']}<br>
  Printed: {now}
</div>
</body></html>"""
    return html


def generate_jobcard_excel_export(status_filter=None, priority_filter=None, technician_id=None, date_from=None, date_to=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Job Cards"

    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="0D6EFD", end_color="0D6EFD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    headers = ["Job #", "Title", "Customer", "Technician", "Status", "Priority",
               "Requested", "Scheduled", "Completed", "Due Date",
               "Est. Hours", "Actual Hours", "Parts Cost", "Labour Cost", "Total Cost",
               "Amount Charged", "Signed Off", "Customer PO", "Created By"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    query = db_session.query(JobCard)
    if status_filter:
        query = query.filter(JobCard.status == JobCardStatus[status_filter])
    if priority_filter:
        query = query.filter(JobCard.priority == Priority[priority_filter])
    if technician_id:
        query = query.filter(JobCard.technician_id == int(technician_id))
    if date_from:
        query = query.filter(JobCard.requested_date >= datetime.strptime(date_from, "%Y-%m-%d").date())
    if date_to:
        query = query.filter(JobCard.requested_date <= datetime.strptime(date_to, "%Y-%m-%d").date())

    jobcards = query.order_by(JobCard.created_at.desc()).all()

    for row, jc in enumerate(jobcards, 2):
        customer = db_session.query(Customer).get(jc.customer_id)
        tech = db_session.query(Technician).get(jc.technician_id) if jc.technician_id else None
        data = [
            jc.job_number, jc.title,
            customer.name if customer else '',
            tech.name if tech else '',
            jc.status.value.replace('_', ' ').title(),
            jc.priority.value.upper(),
            jc.requested_date.strftime("%Y-%m-%d") if jc.requested_date else '',
            jc.scheduled_date.strftime("%Y-%m-%d") if jc.scheduled_date else '',
            jc.completed_date.strftime("%Y-%m-%d") if jc.completed_date else '',
            jc.due_date.strftime("%Y-%m-%d") if jc.due_date else '',
            jc.estimated_hours, jc.actual_hours,
            round(jc.total_parts_cost, 2), round(jc.total_labour_cost, 2),
            round(jc.total_cost, 2), round(jc.amount_charged, 2),
            'Yes' if jc.signed_off else 'No',
            jc.customer_po or '', jc.created_by or ''
        ]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border

    widths = [15, 30, 20, 20, 15, 10, 12, 12, 12, 12, 10, 10, 12, 12, 12, 12, 10, 12, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
